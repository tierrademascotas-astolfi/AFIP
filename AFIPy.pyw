# -------------------------------------------------------------------------------------------------
# Paquetes.
# -------------------------------------------------------------------------------------------------

# 1. Establecer un registro de días y facturación en ese día.
# 2. Cuando se clickea en el programa, automatiza los días según el último registro -5 si el último 
# día es muy lejano.
# 3. Acción de preguntar si las fechas son correctas por seguridad.

# -------------------------------------------------------------------------------------------------
# Paquetes.
# -------------------------------------------------------------------------------------------------

import os
import time
import shutil
import openpyxl
import pandas as pd
from typing import Any, List, Tuple
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment
from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import tkinter as tk
from tkinter import messagebox, simpledialog


# -------------------------------------------------------------------------------------------------
# Listas.
# -------------------------------------------------------------------------------------------------

Tipos_de_Comprobantes = {'Factura C': 2,
                         'Nota de Débito C': 3,
                         'Nota de Crédito C': 4,
                         'Recibo C': 5,
                         'Factura de Crédito Electrónica MiPyMEs (FCE) C': 120,
                         'Nota de Débito Electrónica MiPyMEs (FCE) C': 121,
                         'Nota de Crédito Electrónica MiPyMEs (FCE) C': 122}

Tipos_de_Conceptos = {'Productos': 1,
                      'Servicios': 2,
                      'Productos y Servicios': 3}

Tipos_Condicion_IVA = {"IVA Responsable Inscripto": "1",
                       "IVA Sujeto Exento": "4",
                       "Consumidor Final": "5",
                       "Responsable Monotributo": "6",
                       "Sujeto No Categorizado": "7",
                       "Proveedor del Exterior": "8",
                       "Cliente del Exterior": "9",
                       "IVA Liberado - Ley Nº 19.640": "10",
                       "Monotributista Social": "13",
                       "IVA No Alcanzado": "15",
                       "Monotributista Trabajador Independiente Promovido": "16"}


# -------------------------------------------------------------------------------------------------
# Variables globales.
# -------------------------------------------------------------------------------------------------

Ruta_Descarga_Payway = 'C:/Users/tomas/Downloads'
Nombre_Viejo = 'Movimientos En Linea en pesos Delimitado por comas.csv'
Ruta_Nueva_Payway = 'G:/Mi unidad/AFIP/Tablas' 
Nombre_Nuevo = 'Payway.csv'
Email_Payway = 'carolina8101924@gmail.com'
Contraseña_Payway = '123Nogue$'
Usuario_AFIP = '27202147025'
Contraseña_AFIP = '123Patricio$'
Empresa = 'MARQUEZ CAROLINA MARIEL'
Punto_de_Ventas = '00002-Las Piedras 2837 - Kilometro 45, Buenos Aires'
Tipo_Comprobante = 'Factura C'
Tipo_Concepto = 'Productos'
Tipo_Condicion_IVA = 'Consumidor Final'
Tipo_Pago = 'Contado'
Tipo_Comprobante_Valor = Tipos_de_Comprobantes.get(Tipo_Comprobante)
Tipo_Concepto_Valor = Tipos_de_Conceptos.get(Tipo_Concepto)
Tipo_Condicion_IVA_Valor = Tipos_Condicion_IVA.get(Tipo_Condicion_IVA)
RUTA_ICONO = 'G:/Mi unidad/AFIP/Icon.ico'


# -------------------------------------------------------------------------------------------------
# Funciones.
# -------------------------------------------------------------------------------------------------

# Etapa 0. Fechas.

def Generar_Lista_Dias_Previos() -> List[str]:

    """
    Genera una lista de fechas basada en el día actual de la semana.
    Dependiendo del día de hoy, incluye días específicos previos hasta hoy
    en formato DD/MM/AA.

    Retorna:
        Una lista de cadenas representando las fechas especificadas en formato DD/MM/AA.

    Ejemplo:
        Si hoy es Lunes (01/01/24):
        >>> Generar_Lista_Dias_Previos()
        ['28/12/23', '29/12/23', '30/12/23', '31/12/23', '01/01/24']

        Si hoy es Jueves (04/01/24):
        >>> Generar_Lista_Dias_Previos()
        ['01/01/24', '02/01/24', '03/01/24', '04/01/24']
    """

    # Detectar fecha actual y día de la semana.
    Hoy = datetime.today()
    Dia_Semana = Hoy.weekday()  # 0 = Lunes, ..., 6 = Domingo.

    # Definir el mapeo de días a sus días previos correspondientes.
    Dias_Atras = {
        0: [4, 3, 2, 1, 0],  # Lunes: Jueves, Viernes, Sábado, Domingo, Lunes.
        1: [4, 3, 2, 1, 0, -1],  # Martes: Jueves, Viernes, Sábado, Domingo, Lunes, Martes.
        2: [2, 1, 0],  # Miércoles: Lunes, Martes, Miércoles.
        3: [3, 2, 1, 0],  # Jueves: Lunes, Martes, Miércoles, Jueves.
        4: [4, 3, 2, 1, 0],  # Viernes: Lunes, Martes, Miércoles, Jueves, Viernes.
        5: [2, 1, 0],  # Sábado: Jueves, Viernes, Sábado.
        6: [3, 2, 1, 0],  # Domingo: Jueves, Viernes, Sábado, Domingo.
    }

    # Generar la lista de fechas basada en el día de la semana.
    Fechas = [
        (Hoy - timedelta(days=Dias)).strftime("%d/%m/%Y")
        for Dias in Dias_Atras[Dia_Semana]
    ]

    return Fechas

def Verificar_Fechas(Fechas: List[str]) -> Tuple[bool, List[str]]:

    """
    Muestra un cuadro de diálogo para verificar las fechas y permite editarlas 
    si es necesario.

    Parámetros:
        Fechas: Lista de fechas a verificar.

    Retorna:
        Tupla con un booleano indicando si se confirmaron las fechas y la lista 
        de fechas (original o modificada).
    """

    Root = tk.Tk()
    Root.withdraw()  # Ocultar Root.
    Root.iconbitmap(RUTA_ICONO) 

    # Formatear las fechas para mostrar.
    Primera_Fecha = Fechas[0]
    Ultima_Fecha = Fechas[-1]
    Mensaje = f"Período de facturación:\nDesde: {Primera_Fecha}\n" \
              f"Hasta: {Ultima_Fecha}\n\n¿Son correctas las fechas?"
    
    # Mostrar el cuadro de diálogo de confirmación.
    Confirmado = messagebox.askyesno("Verificación de fechas", Mensaje)
    
    if not Confirmado:
        # Si no se confirman las fechas, permitir edición.
        Nueva_Primera_Fecha = simpledialog.askstring(
            "Editar fechas", 
            "Ingrese la fecha de inicio (DD/MM/YYYY):",
            initialvalue = Primera_Fecha
        )
        
        Nueva_Ultima_Fecha = simpledialog.askstring(
            "Editar fechas", 
            "Ingrese la fecha final (DD/MM/YYYY):",
            initialvalue = Ultima_Fecha
        )
        
        if Nueva_Primera_Fecha and Nueva_Ultima_Fecha:
            try:
                # Convertir las fechas ingresadas a datetime.
                Fecha_Inicio = datetime.strptime(Nueva_Primera_Fecha, "%d/%m/%Y")
                Fecha_Fin = datetime.strptime(Nueva_Ultima_Fecha, "%d/%m/%Y")
                
                # Generar lista de fechas entre inicio y fin.
                Nuevas_Fechas = []
                Fecha_Actual = Fecha_Inicio
                while Fecha_Actual <= Fecha_Fin:
                    Nuevas_Fechas.append(Fecha_Actual.strftime("%d/%m/%Y"))
                    Fecha_Actual += timedelta(days=1)
                
                return True, Nuevas_Fechas
            except ValueError:
                messagebox.showerror("Error", 
                    "Formato de fecha inválido. Se usarán las fechas originales.")
                return True, Fechas
    
    return True, Fechas

# Etapa 1. Payway.

def Descargar_CSV_De_Payway(Navegador: WebDriver, Email: str, 
                           Contraseña: str) -> WebDriver:

    """
    Automatiza el proceso de inicio de sesión y descarga un archivo CSV 
    desde el sitio web de Payway.

    Parámetros:
        Navegador: La instancia de WebDriver de Selenium para controlar el navegador.
        Email: La dirección de correo electrónico usada para el inicio de sesión.
        Contraseña: La contraseña usada para el inicio de sesión.

    Retorna:
        La instancia de WebDriver después de descargar el archivo CSV.

    Ejemplo:
        >>> Navegador = AlgunaInstanciaDeWebDriver()
        >>> Descargar_CSV_De_Payway(Navegador, "test@ejemplo.com", "contraseña123")

    """

    URL_Login = 'https://mi.payway.com.ar/ms/ui-login/login'

    # Abrir la página de inicio de sesión.
    Navegador.get(URL_Login)

    # Esperar hasta que el campo "Email" esté presente.
    WebDriverWait(Navegador, 1000).until(
        EC.presence_of_element_located((By.NAME, "email"))
    )

    # Localizar el campo de email por su nombre.
    Campo_Email = Navegador.find_element(By.NAME, "email")

    # Ingresar el email en el campo.
    Campo_Email.send_keys(Email)

    # Localizar el campo de contraseña por su nombre.
    Campo_Contraseña = Navegador.find_element(By.NAME, "pwd")

    # Ingresar la contraseña en el campo.
    Campo_Contraseña.send_keys(Contraseña)

    # Presionar Enter para enviar el formulario.
    Campo_Contraseña.send_keys(Keys.RETURN)

    # Esperar 30 segundos para asegurar que la página cargue completamente.
    time.sleep(30)

    URL_Movimientos = 'https://mi.payway.com.ar/movimientos/en-linea'

    # Abrir la página de movimientos.
    Navegador.get(URL_Movimientos)

    # Esperar hasta que el botón "Descargar CSV" esté presente.
    Boton_Descarga = WebDriverWait(Navegador, 1000).until(
        EC.presence_of_element_located((By.CLASS_NAME, "sc-crHmcD.dMPykX"))
    )

    # Hacer clic en el botón "Descargar CSV".
    Boton_Descarga.click()

    # Esperar hasta que el botón "CSV separado por comas" esté presente.
    Boton_CSV_Comas = WebDriverWait(Navegador, 1000).until(
        EC.presence_of_element_located(
            (By.XPATH, "//button[@class='sc-gJbFto fMZkYs']//span[text()='Comas']")
        )
    )

    # Hacer clic en el botón "CSV separado por comas".
    Boton_CSV_Comas.click()

    return Navegador

# Etapa 2. Procesamiento del archivo CSV.

def Aplicar_Formato_Excel(Ruta_Archivo: str) -> None:

    """
    Aplica formato específico a un archivo Excel usando OpenPyXL.

    Parámetros:
        Ruta_Archivo: La ruta al archivo Excel a formatear.

    Retorna:
        None

    Ejemplo:
        >>> Aplicar_Formato_Excel('salida.xlsx')

    """

    # Verificar si el archivo existe.
    if not os.path.isfile(Ruta_Archivo):
        raise FileNotFoundError(f"El archivo '{Ruta_Archivo}' no existe.")
    
    # Cargar el libro de trabajo y seleccionar la hoja activa.
    Libro_Trabajo = load_workbook(Ruta_Archivo)

    # Verificar que el libro tenga hojas.
    if not Libro_Trabajo.sheetnames:
        raise ValueError("No se encontraron hojas en el libro de trabajo.")
    Hoja = Libro_Trabajo.active

    # Verificar que la hoja tenga contenido.
    if Hoja.max_row == 0 or Hoja.max_column == 0: # type: ignore
        raise ValueError("La hoja activa está vacía y no tiene contenido para formatear.")

    # Establecer anchos de columna.
    Hoja.column_dimensions['A'].width = 15 # type: ignore
    Hoja.column_dimensions['B'].width = 40 # type: ignore
    Hoja.column_dimensions['C'].width = 15 # type: ignore

    # Aplicar estilo general sin decimales para la columna C ("Precio").
    Estilo_General = NamedStyle(name="general")
    Estilo_General.number_format = '0'

    if 'C' in Hoja.column_dimensions: # type: ignore
        for Celda in Hoja['C']: # type: ignore
            Celda.style = Estilo_General

    # Centrar todas las celdas.
    Alineacion_Centro = Alignment(horizontal='center', vertical='center')

    for Columna in Hoja.iter_cols():  # type: ignore
        # Usar iter_cols para iterar de manera confiable sobre las columnas
        for Celda in Columna:
            Celda.alignment = Alineacion_Centro

    # Guardar el libro de trabajo formateado.
    Libro_Trabajo.save(Ruta_Archivo)

def Mover_Y_Renombrar_Archivo(Ruta_Original: str, Nombre_Original: str, 
                            Ruta_Nueva: str, Nombre_Nuevo: str) -> None:
   
   """
   Mueve un archivo desde su ubicación original a una nueva ubicación 
   especificada y lo renombra en el proceso.

   Parámetros:
       Ruta_Original: La ruta del directorio que contiene el archivo original.
       Nombre_Original: El nombre del archivo a mover.
       Ruta_Nueva: La ruta del nuevo directorio donde se moverá el archivo.
       Nombre_Nuevo: El nuevo nombre a asignar al archivo después de moverlo.

   Retorna:
       None

   Ejemplo:
       >>> Mover_Y_Renombrar_Archivo(
       ...     "C:/original", 
       ...     "archivo.txt", 
       ...     "C:/nueva/ubicacion", 
       ...     "nuevo_nombre_archivo.txt"
       ... )

   """

   # Combinar la ruta y nombre original para formar la ruta completa original.
   Ruta_Original_Completa = os.path.join(Ruta_Original, Nombre_Original)

   # Combinar la nueva ruta y nuevo nombre para formar la ruta destino completa.
   Ruta_Nueva_Completa = os.path.join(Ruta_Nueva, Nombre_Nuevo)

   # Usar shutil para mover y renombrar el archivo a la nueva ubicación.
   shutil.move(Ruta_Original_Completa, Ruta_Nueva_Completa)

def Procesar_Y_Guardar_Dataframe(Dataframe: pd.DataFrame, Ruta_Salida: str):

    """
    Procesa un DataFrame de Pandas, aplica formato y lo guarda como archivo Excel.

    Parámetros:
        Dataframe: El DataFrame a procesar y guardar.
        Ruta_Salida: La ruta del archivo donde guardar el Excel.

    Retorna:
        None

    Ejemplo:
        >>> df = pd.DataFrame({'Fecha': ['02/04/2024'], 'Precio': [100]})
        >>> Procesar_Y_Guardar_Dataframe(df, 'salida.xlsx')

    """

    # Columnas.
    Orden_Columnas = ['Fecha', 'Descripción', 'Precio']

    # Crear un nuevo DataFrame.
    df = pd.DataFrame(columns = Orden_Columnas)

    # Agregar las columnas 'Descripción' y 'Precio' del DataFrame original.
    df['Precio'] = Dataframe['MONTO_BRUTO']
    df['Fecha'] = Dataframe['FECHA']

    # DataFrame con precios.
    Sistema = 'G:/Mi unidad/Tablas y datos/Exportar.xls'
    df_Sistema = pd.read_excel(Sistema)

    def Asignar_Descripcion(DataFrame: pd.DataFrame, Precio: float) -> str:

        """
        Asigna una descripción del dataframe basada en el precio más cercano.

        Este método calcula la diferencia absoluta entre el precio proporcionado 
        y los precios en la columna 'Precio' del dataframe. Luego selecciona la 
        descripción de la columna 'Descripción' correspondiente al precio más cercano.

        Parámetros:
        - DataFrame (pd.DataFrame): Un DataFrame de pandas que debe contener las 
          columnas 'Precio' y 'Descripción'.
        - Precio (float): El precio objetivo para encontrar la coincidencia más cercana.

        Retorna:
        - str: La descripción asociada con el precio más cercano encontrado.

        """

        # Encuentra el índice del precio más cercano.
        Indice_Cercano = (DataFrame['Precio'] - Precio).abs().idxmin()  

        # Retorna la descripción correspondiente.
        return str(DataFrame.loc[Indice_Cercano, 'Descripción'])

    # Aplicar la función al DataFrame.
    df['Descripción'] = df.apply(lambda Fila: Asignar_Descripcion(df_Sistema, 
                                                                 Fila['Precio']), 
                                axis=1)

    # Convertir columna 'Fecha' a datetime y ordenar por fecha.
    df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y')
    df = df.sort_values(by='Fecha')

    # Formatear columna 'Fecha' de vuelta al formato 'dd/mm/yyyy'.
    df['Fecha'] = df['Fecha'].dt.strftime('%d/%m/%Y')

    # Resetear el índice.
    df = df.reset_index(drop=True)

    # Guardar df a Excel.
    df.to_excel(Ruta_Salida, index=False)

    # Aplicar formato adicional usando OpenPyXL.
    Aplicar_Formato_Excel(Ruta_Salida)

def Dividir_Filas_Por_Umbral(Data_Frame: pd.DataFrame, Columna: str, 
                           Umbral: int) -> pd.DataFrame:

   """
   Divide las filas en el DataFrame cuando el valor en la columna especificada 
   supera el umbral establecido. Si un valor excede el umbral, se divide el valor 
   por la mitad del umbral y crea nuevas filas con ese valor.

   Parámetros:
       Data_Frame (pd.DataFrame): El DataFrame de entrada.
       Columna (str): El nombre de la columna donde se aplicará el umbral.
       Umbral (int): El valor umbral para dividir las filas.

   Retorna:
       pd.DataFrame: Un nuevo DataFrame con las filas modificadas.

   """

   # Crear una lista vacía para almacenar las nuevas filas.
   Nuevas_Filas = []

   # Recorrer cada fila en el DataFrame.
   for _, Fila in Data_Frame.iterrows():
       Valor = Fila[Columna]
       # Si el valor excede el umbral.
       if Valor > Umbral:
           # Calcular cuántas filas crear.
           Valor_Division = Umbral // 2
           Cantidad_Filas = int(Valor // Valor_Division)
           Valor_Restante = Valor % Valor_Division

           # Crear nuevas filas con los valores divididos.
           for _ in range(Cantidad_Filas):
               Nueva_Fila = Fila.copy()
               Nueva_Fila[Columna] = Valor_Division
               Nuevas_Filas.append(Nueva_Fila)

           # Agregar el resto como una fila extra si existe.
           if Valor_Restante > 0:
               Nueva_Fila = Fila.copy()
               Nueva_Fila[Columna] = Valor_Restante
               Nuevas_Filas.append(Nueva_Fila)
       else:
           # Si el valor está por debajo del umbral, mantener la fila original.
           Nuevas_Filas.append(Fila)

   # Retornar el nuevo DataFrame.
   return pd.DataFrame(Nuevas_Filas)

# Etapa 3. AFIP.

def Inicializar_Navegador_Chrome() -> webdriver.Chrome:

   """
   Inicializa una instancia de Chrome WebDriver con opciones específicas.

   Retorna:
       Una instancia de Chrome WebDriver.

   Ejemplo:
       >>> navegador = Inicializar_Navegador_Chrome()

   """

   Opciones = webdriver.ChromeOptions()
   Opciones.add_argument('--ignore-certificate-errors')
   Opciones.add_argument('--ignore-ssl-errors')

   Navegador = webdriver.Chrome(options=Opciones)
   Navegador.maximize_window()
   return Navegador

def Esperar_Descarga(Ruta_Archivo: str, Tiempo_Limite: int = 30, 
                     Intervalo: int = 2) -> None:

    """
    Espera a que un archivo se descargue en la ruta especificada.

    Parámetros:
        Ruta_Archivo: La ruta al archivo que se espera descargar.
        Tiempo_Limite: Tiempo máximo de espera en segundos.
        Intervalo: Tiempo entre verificaciones de existencia del archivo en segundos.

    Retorna:
        None

    Ejemplo:
        >>> Esperar_Descarga("/ruta/al/archivo.csv")

    """

    Tiempo_Inicio = time.time()
    while not os.path.exists(Ruta_Archivo):
        if time.time() - Tiempo_Inicio > Tiempo_Limite:
            raise TimeoutError("Se excedió el tiempo límite de descarga.")
        time.sleep(Intervalo)

def Iniciar_Sesion_AFIP(Navegador: webdriver.Chrome, Usuario: str, 
                       Contraseña: str, Empresa: str) -> webdriver.Chrome:

   """
   Inicia sesión en la plataforma AFIP.

   Parámetros:
       Navegador: La instancia de Selenium WebDriver.
       Usuario: El nombre de usuario de AFIP.
       Contraseña: La contraseña de AFIP.
       Empresa: La empresa a seleccionar.

   Retorna:
       La instancia de WebDriver después del inicio de sesión.

   Ejemplo:
       >>> Iniciar_Sesion_AFIP(navegador, "usuario@ejemplo.com", 
       ...                     "contraseña123", "Mi Empresa")

   """

   URL_Login = 'https://auth.afip.gob.ar/contribuyente_/login.xhtml'
   Navegador.get(URL_Login)

   WebDriverWait(Navegador, 10).until(
       EC.presence_of_element_located((By.ID, 'F1:username'))
   )

   Campo_Usuario = Navegador.find_element(By.ID, 'F1:username')
   Campo_Usuario.send_keys(Usuario)
   Campo_Usuario.send_keys(Keys.RETURN)

   WebDriverWait(Navegador, 10).until(
       EC.presence_of_element_located((By.ID, 'F1:password'))
   )

   Campo_Contraseña = Navegador.find_element(By.ID, 'F1:password')
   Campo_Contraseña.send_keys(Contraseña)
   Campo_Contraseña.send_keys(Keys.RETURN)

   # Navegar a la sección de facturas.
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located(
       (By.CSS_SELECTOR, 'h3.roboto-font.regular.p-y-0.m-y-0.h4')))
   Enlace_Facturas = Navegador.find_element(By.CSS_SELECTOR, 
       'h3.roboto-font.regular.p-y-0.m-y-0.h4')
   Enlace_Facturas.click()

   WebDriverWait(Navegador, 10).until(EC.number_of_windows_to_be(2))
   Ventanas = Navegador.window_handles
   Navegador.switch_to.window(Ventanas[1])

   XPath_Empresa = f"//input[@value='{Empresa}']"  
   Boton_Empresa = Navegador.find_element(By.XPATH, XPath_Empresa)
   Boton_Empresa.click()

   return Navegador

def Generar_Factura(Navegador: webdriver.Chrome, Punto_Venta: str, 
                   Tipo_Factura: int, Fecha: str, Tipo_Concepto: int, 
                   Tipo_IVA: int, Tipo_Pago: str, Descripcion: str, 
                   Precio: float) -> webdriver.Chrome:

   """
   Genera una factura utilizando el sistema en línea de AFIP.

   Parámetros:
       Navegador: Instancia de Chrome WebDriver.
       Punto_Venta: Identificador del punto de venta.
       Tipo_Factura: Tipo de factura a generar.
       Fecha: Fecha de la factura en formato "YYYY-MM-DD".
       Tipo_Concepto: Tipo de concepto para la factura.
       Tipo_IVA: Condición de IVA del receptor.
       Tipo_Pago: Tipo de pago (ej: 'Contado').
       Descripcion: Descripción del producto/servicio.
       Precio: Precio del producto/servicio.

   Retorna:
       La instancia de WebDriver después de completar el proceso.

   Ejemplo:
       >>> Generar_Factura(Navegador, "0001", 2, "2024-12-26", 2, 1, 
       ...                 "Contado", "Servicio", 1000.0)

   """

   # Paso 1: Generar Factura
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located(
       (By.XPATH, "//span[@class='ui-button-text' and text()='Generar Comprobantes']")))
   Boton_Generar = Navegador.find_element(By.XPATH, 
       "//span[@class='ui-button-text' and text()='Generar Comprobantes']")
   Boton_Generar.click()
   time.sleep(1)

   # Paso 2: Seleccionar Punto de Venta y Tipo de Factura
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located((By.ID, "puntodeventa")))
   Lista_Punto_Venta = Navegador.find_element(By.ID, "puntodeventa")
   Opcion_Punto_Venta = Lista_Punto_Venta.find_element(
       By.XPATH, f"//option[contains(text(), '{Punto_Venta}')]")
   Opcion_Punto_Venta.click()

   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located(
       (By.ID, "universocomprobante")))
   Lista_Tipo_Factura = Navegador.find_element(By.ID, "universocomprobante")
   time.sleep(1)
   Select(Lista_Tipo_Factura).select_by_value(str(Tipo_Factura))

   Boton_Continuar = Navegador.find_element(By.XPATH, "//input[@value='Continuar >']")
   Boton_Continuar.click()

   # Paso 3: Ingresar Fecha y Concepto
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located((By.ID, "fc")))
   Campo_Fecha = Navegador.find_element(By.ID, "fc")
   Campo_Fecha.clear()
   time.sleep(1)
   Campo_Fecha.send_keys(Fecha)

   Lista_Concepto = Navegador.find_element(By.ID, "idconcepto")
   time.sleep(1)
   Select(Lista_Concepto).select_by_value(str(Tipo_Concepto))
   Boton_Continuar = Navegador.find_element(By.XPATH, "//input[@value='Continuar >']")
   Boton_Continuar.click()

   # Paso 4: Seleccionar Condición IVA y Tipo de Pago
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located((By.ID, "idivareceptor")))
   Lista_IVA = Navegador.find_element(By.ID, "idivareceptor")
   Select(Lista_IVA).select_by_value(str(Tipo_IVA))

   if Tipo_Pago.lower() == 'contado':
       Boton_Pago = Navegador.find_element(By.ID, "formadepago1")
       Boton_Pago.click()

   Boton_Continuar = Navegador.find_element(By.XPATH, "//input[@value='Continuar >']")
   Boton_Continuar.click()

   # Paso 5: Ingresar Descripción del Producto y Precio
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located(
       (By.ID, "detalle_descripcion1")))
   Campo_Descripcion = Navegador.find_element(By.ID, "detalle_descripcion1")
   Campo_Descripcion.clear()
   time.sleep(1)
   Campo_Descripcion.send_keys(Descripcion)

   Campo_Precio = Navegador.find_element(By.ID, "detalle_precio1")
   Campo_Precio.clear()
   time.sleep(1)
   Campo_Precio.send_keys(str(Precio))

   Boton_Continuar = Navegador.find_element(By.XPATH, "//input[@value='Continuar >']")
   Boton_Continuar.click()

   # Paso 6: Confirmar
   WebDriverWait(Navegador, 10).until(EC.presence_of_element_located(
       (By.XPATH, "//input[@value='Confirmar Datos...']")))
   Boton_Confirmar = Navegador.find_element(By.XPATH, 
       "//input[@value='Confirmar Datos...']")
   time.sleep(1)
   Boton_Confirmar.click()

   Alerta = WebDriverWait(Navegador, 10).until(EC.alert_is_present())
   Alerta.accept()

   Boton_Menu = Navegador.find_element(By.XPATH, "//input[@value='Menú Principal']")
   time.sleep(1)
   Boton_Menu.click()

   return Navegador


# Etapa 4. Cierre.

def Cerrar_Todas_Las_Pestanas(Navegador: WebDriver) -> None:

   """
   Cierra todas las pestañas de Chrome abiertas por la instancia de WebDriver.

   Parámetros:
       Navegador: La instancia de Selenium WebDriver que controla el navegador.

   Retorna:
       None

   Ejemplo:
       >>> Cerrar_Todas_Las_Pestanas(navegador)

   """

   # Cerrar todas las pestañas terminando la sesión de WebDriver.
   Navegador.quit()


# -------------------------------------------------------------------------------------------------
# Implementación.
# -------------------------------------------------------------------------------------------------


# Etapa 0. Fechas.

Fechas = Generar_Lista_Dias_Previos()

# Verificación de fechas.
Confirmado, Fechas = Verificar_Fechas(Fechas)

if not Confirmado:
    print("Programa cancelado por el usuario")
    exit()

# Etapa 1. Payway.

# Abrir pestaña de Payway.
Payway = Inicializar_Navegador_Chrome()

# Descargar documento de Payway.
Descargar_CSV_De_Payway(Payway, Email_Payway, Contraseña_Payway)

# Esperar descarga del CSV.
Esperar_Descarga(f'{Ruta_Descarga_Payway}/{Nombre_Viejo}')

# Mover y renombrar archivo CSV.
Mover_Y_Renombrar_Archivo(Ruta_Descarga_Payway, Nombre_Viejo, 
                         Ruta_Nueva_Payway, Nombre_Nuevo)


# Etapa 2. Procesamiento del archivo CSV.

# Crear DataFrame del CSV.
df = pd.read_csv(Ruta_Nueva_Payway + '/' + Nombre_Nuevo, skiprows=1)

# Dividir filas con valores mayores a 100000.
df = Dividir_Filas_Por_Umbral(df, 'MONTO_BRUTO', 100000)

# Ruta del archivo con los precios a subir.
Archivo = f'{Ruta_Nueva_Payway}/AFIP.xlsx'

# Retocar CSV para que quede preparado para su utilización.
Procesar_Y_Guardar_Dataframe(df, Archivo)

# Crear DataFrame final para facturar.
df = pd.read_excel(Archivo)

# Filtrar df con las fechas especificadas al comienzo.
df = df[df['Fecha'].isin(Fechas)]

# Resetear el índice del DataFrame.
df = df.reset_index(drop=True)

# Variables para las columnas.
Fecha = df['Fecha']
Descripcion = df['Descripción']
Precio = df['Precio']


# Etapa 3. AFIP.

# Abrir pestaña de AFIP.
Afip = Inicializar_Navegador_Chrome()

# Loguearse en AFIP.
Iniciar_Sesion_AFIP(Afip, Usuario_AFIP, Contraseña_AFIP, Empresa)

# Bucle de facturación en AFIP producto a producto.
if Tipo_Comprobante_Valor is not None and Tipo_Concepto_Valor is not None and Tipo_Condicion_IVA_Valor is not None:
    for i in range(0, len(Descripcion)):
        Generar_Factura(Afip, 
                       Punto_de_Ventas, 
                       Tipo_Comprobante_Valor,
                       Fecha[i], 
                       Tipo_Concepto_Valor, 
                       int(Tipo_Condicion_IVA_Valor), 
                       Tipo_Pago, 
                       Descripcion[i], 
                       Precio[i])
else:
    raise ValueError("Tipo_Comprobante_Valor es None. Por favor, verificar el valor de Tipo_Comprobante.")


# Etapa 4. Cierre.

# Cerrar todas las pestañas.
#Cerrar_Todas_Las_Pestanas(Afip)
#Cerrar_Todas_Las_Pestanas(Payway)